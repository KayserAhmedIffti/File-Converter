import pandas as pd
import os
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, Label, Entry, Button, filedialog
from tkinter import ttk
from tkcalendar import DateEntry
from copy import copy

def process_data(assignee_name, report_start, report_end, file_path):
    try:
        if getattr(sys, 'frozen', False):
            curdir = sys._MEIPASS
        else:
            curdir = os.path.dirname(os.path.abspath(__file__))
        
        print(f"Current directory: {curdir}")  # Debugging statement

        # Set the working directory to the location of the executable or script
        os.chdir(curdir)

        df = pd.read_excel(file_path)
        output_directory = os.path.join(curdir, 'Readable_Weekly_Task')
        os.makedirs(output_directory, exist_ok=True)
        
        print(f"Output directory: {output_directory}")  # Debugging statement
        
        df['Due Date'] = pd.to_datetime(df['Due Date'] / 1000, unit='s')
        start_date = pd.to_datetime(report_start)
        end_date = pd.to_datetime(report_end)
        
        filtered_df = df[(df['Due Date'] >= start_date) & (df['Due Date'] <= end_date)]
        filtered_df = filtered_df[filtered_df['Assignees'].str.contains(assignee_name)]
        
        grouped = filtered_df.groupby('Folder Name')
        selected_columns = [' Task Name', 'Start Date Text', 'Due Date Text', 'Assignees']
        grouped = grouped[selected_columns]
        
        output_file = os.path.join(output_directory, "redefined_data.xlsx")
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        writer.book.create_sheet('Sheet1')
        
        sheet = writer.book['Sheet1']
        cell = sheet.cell(row=1, column=1, value=f"Weekly Task Report")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        cell.alignment = cell.alignment.copy(wrapText=True, horizontal="center", vertical="center")
        cell.font = cell.font.copy(bold=True, size=18)

        cell = sheet.cell(row=2, column=1, value=f"{assignee_name}")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        cell.alignment = cell.alignment.copy(wrapText=True, horizontal="center", vertical="center")
        cell.font = cell.font.copy(bold=True, size=18)

        cell = sheet.cell(row=3, column=1, value=f"Reporting Dates: {report_start} to {report_end}")
        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
        cell.alignment = cell.alignment.copy(wrapText=True, horizontal="center", vertical="center")
        cell.font = cell.font.copy(bold=True, size=14)

        for folder_name, group in grouped:
            if not group.empty:
                group = group.rename(columns={'Start Date Text': 'Start Time', 'Due Date Text': 'End Time'})
                group['Start Time'] = pd.to_datetime(group['Start Time'], format='%m/%d/%Y, %I:%M:%S %p GMT+6').dt.strftime('%d/%m/%Y, %H:%M')
                group['End Time'] = pd.to_datetime(group['End Time'], format='%m/%d/%Y, %I:%M:%S %p GMT+6').dt.strftime('%d/%m/%Y, %H:%M')

                group_sorted = group.sort_values(by='End Time', ascending=True)
                start_row = sheet.max_row + 3 if sheet.max_row is not None else 0
                group_sorted.to_excel(writer, sheet_name='Sheet1', startrow=start_row, index=False)

                sheet.insert_rows(start_row)
                cell = sheet.cell(row=start_row, column=1, value=f"{folder_name}")
                cell.font = cell.font.copy(bold=True, size=14)

                last_column = sheet.max_column
                sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=last_column)
                cell.alignment = cell.alignment.copy(horizontal="left", vertical="center")
        
        writer.close()
        wb = load_workbook(output_file)
        ws = wb['Sheet1']
        
        column_widths = {'A': 37, 'B': 16, 'C': 16, 'D': 20}
        for column, width in column_widths.items():
            ws.column_dimensions[column].width = width
        ws.row_dimensions[1].height = 21
        ws.row_dimensions[2].height = 21
        ws.row_dimensions[3].height = 17

        for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = cell.alignment.copy(wrapText=True, vertical="center")

        from openpyxl.worksheet.page import PageMargins
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.5, footer=0.5)
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        wb.save(output_file)
        
        print("File saved successfully.")  # Debugging statement

        print("Individual attendance files have been created with additional modifications.")
    
    except PermissionError as e:
        print("PermissionError:", e)
    except Exception as e:
        print("Error:", e)

def browse_file():
    filename = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    file_path_entry.delete(0, 'end')
    file_path_entry.insert(0, filename)

def populate_assignees():
    predefined_assignees = ['Touhidul Islam', 'S M Anwarul Aziz', 'Md Arafin Mahamud', 'Muntasirur Rahman', 'Moue Islam', 'Sumaiya Sabur']
    assignee_combobox['values'] = predefined_assignees

def run():
    assignee_name = assignee_combobox.get()
    report_start = start_date_entry.get_date().strftime('%Y-%m-%d')
    report_end = end_date_entry.get_date().strftime('%Y-%m-%d')
    file_path = file_path_entry.get()
    process_data(assignee_name, report_start, report_end, file_path)

root = Tk()
root.title("ClickUp Data Processor")
Label(root, text="Assignee Name").grid(row=0, column=0, padx=10, pady=10)
assignee_combobox = ttk.Combobox(root)
assignee_combobox.grid(row=0, column=1, padx=10, pady=10)
populate_assignees()  # Populate the combobox with predefined assignees at startup

Label(root, text="Report Start Date (YYYY-MM-DD)").grid(row=1, column=0, padx=10, pady=10)
start_date_entry = DateEntry(root, date_pattern='yyyy-mm-dd')
start_date_entry.grid(row=1, column=1, padx=10, pady=10)

Label(root, text="Report End Date (YYYY-MM-DD)").grid(row=2, column=0, padx=10, pady=10)
end_date_entry = DateEntry(root, date_pattern='yyyy-mm-dd')
end_date_entry.set_date(datetime.today())
end_date_entry.grid(row=2, column=1, padx=10, pady=10)

Label(root, text="Excel File Path").grid(row=3, column=0, padx=10, pady=10)
file_path_entry = Entry(root, width=50)
file_path_entry.grid(row=3, column=1, padx=10, pady=10)
browse_button = Button(root, text="Browse", command=browse_file)
browse_button.grid(row=3, column=2, padx=10, pady=10)
run_button = Button(root, text="Run", command=run)
run_button.grid(row=4, column=1, padx=10, pady=10)
root.mainloop()
